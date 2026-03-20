import pandas as pd
import numpy as np
import os
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl import load_workbook

def create_checklist_pivot(OUTPUT_FILE, CHECKLIST_FILE):
    try:
        MERCHANT_PAYOUT_FILE = r'input_files\merchant_payout_report.csv'

        if not os.path.exists(OUTPUT_FILE):
            print(f"❌ Error: {OUTPUT_FILE} not found.")
            return

        df = pd.read_excel(OUTPUT_FILE)

        # --- Direction Mapping ---
        df['direction_name'] = df['direction'].map({1: 'payout', 2: 'Return', '1': 'payout', '2': 'Return'})
        df['direction_name'] = df['direction_name'].fillna('Unknown')
        df['direction_name'] = pd.Categorical(df['direction_name'], categories=['payout','Return','Unknown'], ordered=True)

        # --- SHEET1 PIVOT ---
        sum_cols = ['GMV','merchant_payable','pg_payable2','cod_payable2','Commission','Tax','cust_shipping_reversal','partial_shipping_rev','pf','product_gst','TCS','TDS','Seller Payable','Diff','shipping_amount2','mp_shipping','additional_delivery_charges2','cart_conv_fee2']
        available_cols = [c for c in sum_cols if c in df.columns]
        pivot_df = pd.pivot_table(df, index='direction_name', values=available_cols, aggfunc='sum')
        pivot_df = pivot_df.reindex(columns=available_cols).reset_index()

        # --- TABLE 1: merchant_payable ---
        pivot_sheet2_top = pd.pivot_table(df, index='direction_name', values=['merchant_payable','pg_payable2','cod_payable2'], aggfunc='sum').reset_index()
        pivot_sheet2_top['Difference'] = pivot_sheet2_top['merchant_payable'] - pivot_sheet2_top['pg_payable2'] - pivot_sheet2_top['cod_payable2']
        pivot_sheet2_top = pivot_sheet2_top[['direction_name','merchant_payable','pg_payable2','cod_payable2','Difference']]

        # --- TABLE 2: Payout Calculation ---
        cols_needed = ['GMV','Commission','Tax','cust_shipping_reversal','partial_shipping_rev','pf','product_gst','TCS','TDS','Seller Payable','merchant_payable']
        for c in cols_needed: 
            if c not in df.columns: df[c] = 0
        pivot_sheet2_bottom = pd.pivot_table(df, index='direction_name', values=cols_needed, aggfunc='sum').reset_index()
        pivot_sheet2_bottom.rename(columns={'partial_shipping_rev':'partial_shipping_reversal'}, inplace=True)
        pivot_sheet2_bottom['Diff'] = pivot_sheet2_bottom['Seller Payable'] - pivot_sheet2_bottom['merchant_payable']
        pivot_sheet2_bottom = pivot_sheet2_bottom[['direction_name','GMV','Commission','Tax','cust_shipping_reversal','partial_shipping_reversal','pf','product_gst','TCS','TDS','Seller Payable','merchant_payable','Diff']]

        # --- TABLE 3: File Wise Comparison ---
        # 1. Data from Compact Revenue (Filtered for Payout only)
        rev_pg = df[df['direction_name'] == 'payout']['pg_payable2'].sum()
        rev_cod = df[df['direction_name'] == 'payout']['cod_payable2'].sum()
        
        # 2. Data from Merchant Payout Report
        m_pg, m_cod = 0, 0
        if os.path.exists(MERCHANT_PAYOUT_FILE):
            df_m = pd.read_csv(MERCHANT_PAYOUT_FILE)
            m_pg = df_m['pg_payable'].sum() if 'pg_payable' in df_m.columns else 0
            m_cod = df_m['cod_payable'].sum() if 'cod_payable' in df_m.columns else 0
        
        comparison_data = [
            ["Compact revenue report", rev_pg, rev_cod, rev_pg + rev_cod],
            ["merchant payout reprot", m_pg, m_cod, m_pg + m_cod]
        ]
        pivot_sheet2_file = pd.DataFrame(comparison_data, columns=['File name', 'pg_payable', 'cod_payable', 'total'])

        # --- WRITE EXCEL ---
        with pd.ExcelWriter(CHECKLIST_FILE, engine='openpyxl') as writer:
            pivot_df.to_excel(writer, sheet_name='Sheet1', index=False)
            sheet_name = "payout calculation"
            
            # Title & Table 1
            pd.DataFrame({"A": ["1  Check list _merchant_payable"]}).to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            pivot_sheet2_top.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
            
            # Title & Table 2
            title2_row = len(pivot_sheet2_top) + 5
            pd.DataFrame({"A": ["2  Check list _Payout Calculation"]}).to_excel(writer, sheet_name=sheet_name, startrow=title2_row, index=False, header=False)
            pivot_sheet2_bottom.to_excel(writer, sheet_name=sheet_name, startrow=title2_row + 2, index=False)

            # Title & Table 3 (The Correction)
            title3_row = title2_row + len(pivot_sheet2_bottom) + 5
            pd.DataFrame({"A": ["3 Check list _File Wise Comparison"]}).to_excel(writer, sheet_name=sheet_name, startrow=title3_row, index=False, header=False)
            pivot_sheet2_file.to_excel(writer, sheet_name=sheet_name, startrow=title3_row + 2, index=False)

            pd.DataFrame().to_excel(writer, sheet_name='Sheet3', index=False)
            pd.DataFrame().to_excel(writer, sheet_name='Sheet4', index=False)

        # --- APPLY FORMATTING ---
        wb = load_workbook(CHECKLIST_FILE)
        ws = wb["payout calculation"]
        header_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
        header_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        def format_block(start_row, data_df):
            end_row = start_row + len(data_df)
            for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=len(data_df.columns)):
                for cell in row:
                    cell.border = border
                    if cell.row == start_row:
                        cell.fill = header_fill
                        cell.font = header_font

        format_block(3, pivot_sheet2_top)
        format_block(title2_row + 3, pivot_sheet2_bottom)
        format_block(title3_row + 3, pivot_sheet2_file)

        wb.save(CHECKLIST_FILE)
        print(f"✅ SUCCESS! 3 Tables created in 'payout calculation' at: {CHECKLIST_FILE}")

    except Exception as e:
        print(f"❌ Error: {e}")