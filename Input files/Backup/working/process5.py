import pandas as pd
import numpy as np
import os
import smtplib
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import load_workbook

# --- CONFIGURATION (Paths) ---
BASE_DIR = r'D:\3) Monthly Task\Payout Calculation\Finance_Automation'

# Input Files
HEADER_FILE = os.path.join(BASE_DIR, 'input_files', 'Header.xlsx')
DATA_FILE = os.path.join(BASE_DIR, 'input_files', 'compact_revenue_report.csv')
HOLD_FILE = os.path.join(BASE_DIR, 'input_files', 'Hold_list.csv')
MM_FILE = os.path.join(BASE_DIR, 'input_files', 'MM.csv')
RECOVERY_FILE = os.path.join(BASE_DIR, 'input_files', 'Recovery.csv')
PERIOD_FILE_PATH = os.path.join(BASE_DIR, 'input_files', 'Period.xlsx')

# Output Files
OUTPUT_FILE = os.path.join(BASE_DIR, 'output_files', 'output.xlsx')
SUMMARY_FILE = os.path.join(BASE_DIR, 'output_files', 'payout_summary.xlsx')
PAYMENT_SHEET_FILE = os.path.join(BASE_DIR, 'output_files', 'payment sheet.xlsx')
JV_WORKING_FILE = os.path.join(BASE_DIR, 'output_files', 'jv_working.xlsx')

# Uploader Templates & Outputs
SAP_TEMPLATE_PATH = os.path.join(BASE_DIR, 'Uploader format', 'Payment_format.xlsx')
JV_FORMAT_FILE = os.path.join(BASE_DIR, 'Uploader format', 'JV_Uploader_format.xlsx')
SAP_OUTPUT_FILE = os.path.join(BASE_DIR, 'output_files', 'sap_payment_uploader.xlsx')
GOOGLE_UPLOADER_FILE = os.path.join(BASE_DIR, 'output_files', 'Google_Payment_Uploader.xlsx')
JV_GROSS_OUTPUT = os.path.join(BASE_DIR, 'output_files', 'jv_uploader_gross.xlsx')
JV_RETURN_OUTPUT = os.path.join(BASE_DIR, 'output_files', 'jv_uploader_return.xlsx')

# Email Credentials
SENDER_EMAIL = "harsh.rawat@paipai.mobi"
SENDER_PASSWORD = "ryklyfoiqfkqyevv"
RECEIVER_EMAIL = ["harsh.rawat@paipai.mobi"]

# ---------------- CORE PROCESSING ---------------- #
def process_revenue_report():
    try:
        header_df = pd.read_excel(HEADER_FILE)
        template_headers = header_df.columns.tolist()
        df = pd.read_csv(DATA_FILE, encoding='latin1', low_memory=False)
        df = df[~df['merchant_id'].isin([1530540, 1530543, 1530541, 870346])]

        def to_num(cols, target_df=df):
            for col in cols:
                if col not in target_df.columns: target_df[col] = 0.0
                target_df[col] = pd.to_numeric(target_df[col], errors='coerce').fillna(0)

        # MID_WID Logic
        df['MID_WID'] = np.where((df['finance_key_type'].fillna('') == 'default') | (df['finance_key_type'].isna()), 
                                 df['merchant_id'].astype(str), 
                                 df['merchant_id'].astype(str) + "_" + df['warehouse_id'].fillna('').astype(str))
        
        to_num(['price', 'qty_ordered', 'direction'])
        base_gmv = (df['price'] * df['qty_ordered']).round(2)
        df['GMV'] = np.where(df['direction'] == 1, base_gmv, np.where(df['direction'] == 2, -base_gmv, 0))

        # Core Calculations
        comm_cols = ['pg_commission', 'mp_commission', 'logistics_penalty', 'reverse_logistics_penalty', 'merchant_cancel_mp_penalty', 'logistics', 'marketing_fee', 'closing_fee', 'deal_setup_fees', 'pf_taxes_comm', 'pf_pac_comm', 'pf_scf_comm']
        to_num(comm_cols + ['cust_shipping_reversal', 'partial_shipping_rev', 'pf_tax', 'pf_packing', 'pf_seller_convenience', 'product_igst', 'product_cgst', 'product_sgst', 'tds_ecom', 'cart_conv_fee'])
        
        df['Commission'] = df[comm_cols].sum(axis=1).round(2)
        df['Tax'] = (df['Commission'] * 0.18).round(2)
        df['pf'] = ((df['pf_tax'] + df['pf_packing'] + df['pf_seller_convenience']) * -1).round(2)
        df['product_gst'] = (df['product_igst'] + df['product_cgst'] + df['product_sgst']).round(2)
        
        tcs_cols = ['tcs_cgst', 'tcs_igst', 'tcs_sgst', 'shipping_tcs_cgst', 'shipping_tcs_igst', 'shipping_tcs_sgst']
        to_num([c for c in tcs_cols if c in df.columns])
        df['TCS'] = df[[c for c in tcs_cols if c in df.columns]].sum(axis=1).round(2)
        
        df['TDS'] = df['tds_ecom'].round(2)
        df['cart_conv_fee2'] = df['cart_conv_fee']

        # Extra columns needed for JV Working
        extra_cols = ['MID_WID', 'direction', 'GMV', 'Commission', 'Tax', 'pf', 'product_gst', 'TCS', 'TDS', 'cart_conv_fee2', 'cust_shipping_reversal']
        
        final_export_cols = list(dict.fromkeys(template_headers + extra_cols))
        df[[c for c in final_export_cols if c in df.columns]].to_excel(OUTPUT_FILE, index=False)
        print("✅ output.xlsx generated successfully.")
        
        generate_jv_working()
    except Exception as e:
        print(f"❌ Process error: {e}")

def generate_jv_working():
    try:
        print("Generating jv_working.xlsx...")
        out_df = pd.read_excel(OUTPUT_FILE)
        out_df['MID_WID'] = out_df['MID_WID'].astype(str)
        
        mm_df = pd.read_csv(MM_FILE, encoding='latin1')
        mm_df.columns = mm_df.columns.str.strip()
        mm_df['MID_WID'] = mm_df['MID_WID'].astype(str).str.strip()
        mm_df = mm_df.drop_duplicates(subset=["MID_WID"], keep="first")

        # 1. GMV Pivot
        jv_pivot = pd.pivot_table(out_df, values='GMV', index='MID_WID', columns='direction', aggfunc='sum').reset_index()
        jv_pivot = jv_pivot.rename(columns={1: 'direction_1', 2: 'direction_2'}).fillna(0)
        if 'direction_1' not in jv_pivot.columns: jv_pivot['direction_1'] = 0
        if 'direction_2' not in jv_pivot.columns: jv_pivot['direction_2'] = 0
        jv_pivot['Total'] = jv_pivot['direction_1'] + jv_pivot['direction_2']

        # 2. Aggregation for 8 extra columns
        extra_cols = ['Commission', 'Tax', 'pf', 'product_gst', 'TCS', 'TDS', 'cart_conv_fee2', 'cust_shipping_reversal']
        existing_extra = [c for c in extra_cols if c in out_df.columns]
        extra_agg = out_df.groupby('MID_WID')[existing_extra].sum().reset_index()

        # 3. Merge and Fix SAP column naming
        final_jv = pd.merge(jv_pivot, mm_df[['MID_WID', 'SAP_New_Code']], on='MID_WID', how='left')
        final_jv = pd.merge(final_jv, extra_agg, on='MID_WID', how='left')
        
        # Rename SAP_New_Code to SAP for the uploader functions
        final_jv = final_jv.rename(columns={'SAP_New_Code': 'SAP'})

        # Final column selection
        final_cols = ['MID_WID', 'SAP', 'direction_1', 'direction_2', 'Total'] + existing_extra
        final_jv[final_cols].to_excel(JV_WORKING_FILE, index=False)
        print("✅ jv_working.xlsx generated successfully.")
    except Exception as e:
        print(f"❌ JV Working Error: {e}")

def generate_jv_uploaders():
    try:
        period_df = pd.read_excel(PERIOD_FILE_PATH, header=None)
        period_val = str(period_df.iloc[0, 0]).strip()
        raw_df = pd.read_csv(DATA_FILE, usecols=['settled_at'], encoding='latin1', nrows=1)
        raw_date = pd.to_datetime(raw_df['settled_at'].iloc[0]).strftime('%d.%m.%Y')
        
        jv_working = pd.read_excel(JV_WORKING_FILE)

        configs = [
            {'output': JV_GROSS_OUTPUT, 'keep': 'sales gross', 'rem': 'sales return', 'col': 'direction_1', 'ser': 1, 'pk_m': '40', 'pk_s': '11', 'acc': '13025016', 'txt': 'MP Payt Gross', 'm': 1},
            {'output': JV_RETURN_OUTPUT, 'keep': 'sales return', 'rem': 'sales gross', 'col': 'direction_2', 'ser': 2, 'pk_m': '50', 'pk_s': '01', 'acc': '13025014', 'txt': 'MP Refunded Gross', 'm': -1}
        ]

        for cfg in configs:
            wb = load_workbook(JV_FORMAT_FILE)
            if cfg['rem'] in wb.sheetnames: del wb[cfg['rem']]
            ws = wb[cfg['keep']]
            
            # Summary Row (Row 2)
            total = round(jv_working[cfg['col']].sum() * cfg['m'], 2)
            sum_data = [cfg['ser'], raw_date, "SA", "MKPL", raw_date, "INR", period_val, period_val, cfg['pk_m'], cfg['acc']]
            for c, v in enumerate(sum_data, 1): ws.cell(row=2, column=c).value = v
            ws.cell(row=2, column=13).value = total

            # Merchant Rows
            curr = 3
            for _, row in jv_working.iterrows():
                val = round(float(row[cfg['col']]) * cfg['m'], 2)
                if val == 0: continue
                ws.cell(row=curr, column=1).value = cfg['ser']
                ws.cell(row=curr, column=9).value = cfg['pk_s']
                ws.cell(row=curr, column=10).value = row["SAP"]
                ws.cell(row=curr, column=13).value = val
                
                # Apply Static Fields to Row 2 and Merchant Rows
                for r in [2, curr]:
                    ws.cell(row=r, column=14).value = "v0"
                    ws.cell(row=r, column=15).value = "1000"
                    ws.cell(row=r, column=16).value = "1000"
                    ws.cell(row=r, column=17).value = "0001"
                    ws.cell(row=r, column=22).value = "Gross"
                    ws.cell(row=r, column=23).value = "FINANCE"
                    ws.cell(row=r, column=27).value = f"{cfg['txt']} {period_val}"
                curr += 1
            wb.save(cfg['output'])
        print("✅ JV Gross and Return uploaders saved.")
    except Exception as e:
        print(f"❌ Uploader Error: {e}")

if __name__ == "__main__":
    process_revenue_report()
    generate_jv_uploaders()