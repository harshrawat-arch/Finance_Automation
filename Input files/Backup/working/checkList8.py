import pandas as pd
import numpy as np
import os
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl import load_workbook

def create_checklist_pivot(OUTPUT_FILE, CHECKLIST_FILE):
    try:

        if not os.path.exists(OUTPUT_FILE):
            print(f"❌ Error: {OUTPUT_FILE} not found.")
            return

        df = pd.read_excel(OUTPUT_FILE)

        # --- Direction Mapping ---
        df['direction_name'] = df['direction'].map(
            {1: 'payout', 2: 'Return', '1': 'payout', '2': 'Return'}
        )
        df['direction_name'] = df['direction_name'].fillna('Unknown')

        # ---------------------------------------------------
        # SHEET1 PIVOT
        # ---------------------------------------------------

        sum_cols = [
            'GMV','merchant_payable','pg_payable2','cod_payable2','Commission','Tax',
            'cust_shipping_reversal','partial_shipping_rev','pf','product_gst','TCS',
            'TDS','Seller Payable','Diff','shipping_amount2','mp_shipping',
            'additional_delivery_charges2','cart_conv_fee2'
        ]

        available_cols = [c for c in sum_cols if c in df.columns]

        pivot_df = pd.pivot_table(
            df,
            index='direction_name',
            values=available_cols,
            aggfunc='sum'
        )

        pivot_df = pivot_df.reindex(columns=available_cols).reset_index()

        # ---------------------------------------------------
        # FIRST PIVOT
        # ---------------------------------------------------

        pivot_sheet2_top = pd.pivot_table(
            df,
            index='direction_name',
            values=['merchant_payable','pg_payable2','cod_payable2'],
            aggfunc='sum'
        ).reset_index()

        pivot_sheet2_top['Difference'] = (
            pivot_sheet2_top['merchant_payable']
            - pivot_sheet2_top['pg_payable2']
            - pivot_sheet2_top['cod_payable2']
        )

        pivot_sheet2_top = pivot_sheet2_top[
            ['direction_name','merchant_payable','pg_payable2','cod_payable2','Difference']
        ]

        # ---------------------------------------------------
        # SECOND PIVOT
        # ---------------------------------------------------

        cols_needed = [
            'GMV','Commission','Tax','cust_shipping_reversal','partial_shipping_rev',
            'pf','product_gst','TCS','TDS','Seller Payable','merchant_payable'
        ]

        for c in cols_needed:
            if c not in df.columns:
                df[c] = 0

        pivot_sheet2_bottom = pd.pivot_table(
            df,
            index='direction_name',
            values=cols_needed,
            aggfunc='sum'
        ).reset_index()

        pivot_sheet2_bottom.rename(
            columns={'partial_shipping_rev':'partial_shipping_reversal'},
            inplace=True
        )

        pivot_sheet2_bottom['Diff'] = (
            pivot_sheet2_bottom['Seller Payable']
            - pivot_sheet2_bottom['merchant_payable']
        )

        pivot_sheet2_bottom = pivot_sheet2_bottom[
            [
                'direction_name','GMV','Commission','Tax',
                'cust_shipping_reversal','partial_shipping_reversal',
                'pf','product_gst','TCS','TDS',
                'Seller Payable','merchant_payable','Diff'
            ]
        ]

        # ---------------------------------------------------
        # WRITE EXCEL
        # ---------------------------------------------------

        with pd.ExcelWriter(CHECKLIST_FILE, engine='openpyxl') as writer:

            # Sheet1 will appear first
            pivot_df.to_excel(writer, sheet_name='Sheet1', index=False)

            sheet_name = "payout calculation"

            title1 = pd.DataFrame({"A": ["1  Check list _merchant_payable"]})
            title1.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

            pivot_sheet2_top.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)

            title2_row = len(pivot_sheet2_top) + 5
            title2 = pd.DataFrame({"A": ["2  Check list _Payout Calculation"]})
            title2.to_excel(writer, sheet_name=sheet_name, startrow=title2_row, index=False, header=False)

            pivot_sheet2_bottom.to_excel(writer, sheet_name=sheet_name, startrow=title2_row + 2, index=False)

            pd.DataFrame().to_excel(writer, sheet_name='Sheet3', index=False)
            pd.DataFrame().to_excel(writer, sheet_name='Sheet4', index=False)

        # ---------------------------------------------------
        # APPLY FORMATTING
        # ---------------------------------------------------

        wb = load_workbook(CHECKLIST_FILE)
        ws = wb["payout calculation"]

        header_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
        header_font = Font(bold=True)

        double_border = Border(
            left=Side(style='double'),
            right=Side(style='double'),
            top=Side(style='double'),
            bottom=Side(style='double')
        )

        for row in ws.iter_rows():
            for cell in row:

                if cell.row in [3, title2_row + 3]:
                    cell.fill = header_fill
                    cell.font = header_font

                if cell.value is not None:
                    cell.border = double_border

        wb.save(CHECKLIST_FILE)

        print("✅ Checklist.xlsx created successfully with formatted tables")

    except Exception as e:
        print(f"❌ Error creating checklist: {e}")