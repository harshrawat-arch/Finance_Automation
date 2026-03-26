import pandas as pd
import numpy as np
import os
import smtplib
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import load_workbook, Workbook

# --- CONFIGURATION (Paths) ---
BASE_DIR = r'D:\3) Monthly Task\Payout Calculation\Finance_Automation'

# Input Files
HEADER_FILE = os.path.join(BASE_DIR, 'input_files', 'Header.xlsx')
DATA_FILE = os.path.join(BASE_DIR, 'input_files', 'compact_revenue_report.csv')
MM_FILE = os.path.join(BASE_DIR, 'input_files', 'MM.csv')
PERIOD_FILE_PATH = os.path.join(BASE_DIR, 'input_files', 'Period.xlsx')
INPUT_PAYMENT_SHEET = os.path.join(BASE_DIR, 'output_files', 'payment sheet.xlsx')

# Template Files
SAP_TEMPLATE_PATH = os.path.join(BASE_DIR, 'Uploader format', 'Payment_format.xlsx')
JV_FORMAT_FILE = os.path.join(BASE_DIR, 'Uploader format', 'JV_Uploader_format.xlsx')

# Output Files
OUTPUT_FILE = os.path.join(BASE_DIR, 'output_files', 'output.xlsx')
JV_WORKING_FILE = os.path.join(BASE_DIR, 'output_files', 'jv_working.xlsx')
SAP_OUTPUT_FILE = os.path.join(BASE_DIR, 'output_files', 'sap_payment_uploader.xlsx')
GOOGLE_UPLOADER_FILE = os.path.join(BASE_DIR, 'output_files', 'Google_Payment_Uploader.xlsx')

# JV Specific Outputs
JV_GROSS_OUTPUT = os.path.join(BASE_DIR, 'output_files', 'jv_uploader_gross.xlsx')
JV_RETURN_OUTPUT = os.path.join(BASE_DIR, 'output_files', 'jv_uploader_return.xlsx')
JV_COMM_OUTPUT = os.path.join(BASE_DIR, 'output_files', 'jv_uploader_commission.xlsx')
JV_CART_OUTPUT = os.path.join(BASE_DIR, 'output_files', 'jv_uploader_cart.xlsx')
FINAL_MERGED_JV = os.path.join(BASE_DIR, 'output_files', 'jv_uploader.xlsx')

# Email Credentials
SENDER_EMAIL = "harsh.rawat@paipai.mobi"
SENDER_PASSWORD = "ryklyfoiqfkqyevv"
RECEIVER_EMAIL = ["harsh.rawat@paipai.mobi"]

# ---------------- 1. REVENUE & WORKING FILE PROCESSING ---------------- #

def process_revenue_report():
    try:
        print("Processing Revenue Report...")
        header_df = pd.read_excel(HEADER_FILE)
        template_headers = header_df.columns.tolist()
        df = pd.read_csv(DATA_FILE, encoding='latin1', low_memory=False)
        
        # Filter Exclusions
        df = df[~df['merchant_id'].isin([1530540, 1530543, 1530541, 870346])]

        def to_num(cols, target_df=df):
            for col in cols:
                if col not in target_df.columns: target_df[col] = 0.0
                target_df[col] = pd.to_numeric(target_df[col], errors='coerce').fillna(0)

        df['MID_WID'] = np.where((df['finance_key_type'].fillna('') == 'default') | (df['finance_key_type'].isna()), 
                                 df['merchant_id'].astype(str), 
                                 df['merchant_id'].astype(str) + "_" + df['warehouse_id'].fillna('').astype(str))
        
        to_num(['price', 'qty_ordered', 'direction', 'shipping_amount', 'cust_shipping_reversal', 'cart_conv_fee'])
        base_gmv = (df['price'] * df['qty_ordered']).round(2)
        df['GMV'] = np.where(df['direction'] == 1, base_gmv, np.where(df['direction'] == 2, -base_gmv, 0))

        comm_cols = ['pg_commission', 'mp_commission', 'logistics_penalty', 'reverse_logistics_penalty', 
                     'merchant_cancel_mp_penalty', 'logistics', 'marketing_fee', 'closing_fee', 
                     'deal_setup_fees', 'pf_taxes_comm', 'pf_pac_comm', 'pf_scf_comm']
        to_num(comm_cols + ['pf_tax', 'pf_packing', 'pf_seller_convenience', 
                            'product_igst', 'product_cgst', 'product_sgst', 'tds_ecom'])
        
        df['Commission'] = df[comm_cols].sum(axis=1).round(2)
        df['Tax'] = (df['Commission'] * 0.18).round(2)
        
        cart_req_cols = ['shipping_amount', 'cust_shipping_reversal', 'cart_conv_fee']
        final_export_cols = list(dict.fromkeys(template_headers + ['MID_WID', 'direction', 'GMV', 'Commission', 'Tax'] + cart_req_cols))
        
        df[[c for c in final_export_cols if c in df.columns]].to_excel(OUTPUT_FILE, index=False)
        
        generate_jv_working(df)
    except Exception as e:
        print(f"❌ Revenue Process error: {e}")

def generate_jv_working(df):
    try:
        mm_df = pd.read_csv(MM_FILE, encoding='latin1')
        mm_df['MID_WID'] = mm_df['MID_WID'].astype(str).str.strip()
        mm_df = mm_df.drop_duplicates(subset=["MID_WID"])

        jv_pivot = pd.pivot_table(df, values='GMV', index='MID_WID', columns='direction', aggfunc='sum').reset_index()
        
        if 1 not in jv_pivot.columns: jv_pivot[1] = 0.0
        if 2 not in jv_pivot.columns: jv_pivot[2] = 0.0
        
        jv_pivot = jv_pivot.rename(columns={1: 'direction_1', 2: 'direction_2'}).fillna(0)

        extra_agg = df.groupby('MID_WID')[['Commission', 'Tax', 'cart_conv_fee']].sum().reset_index()
        final_jv = pd.merge(jv_pivot, mm_df[['MID_WID', 'SAP_New_Code']], on='MID_WID', how='left')
        final_jv = pd.merge(final_jv, extra_agg, on='MID_WID', how='left').rename(columns={'SAP_New_Code': 'SAP'})

        final_jv.to_excel(JV_WORKING_FILE, index=False)
        print("✅ jv_working.xlsx generated.")
    except Exception as e:
        print(f"❌ JV Working Error: {e}")

# ---------------- 2. SAP & GOOGLE UPLOADER GENERATION ---------------- #

def generate_sap_and_google_uploads(period_val, raw_date):
    try:
        print("Generating SAP and Google Uploaders...")
        pay_df = pd.read_excel(INPUT_PAYMENT_SHEET)
        sap_entries = []
        mapping = [('91000467', 'net pg_payable', 'Escrow'), ('91000463', 'net_cod_payable', 'Cod')]

        for gl_code, col, txt_prefix in mapping:
            for _, row in pay_df.iterrows():
                val = pd.to_numeric(row[col], errors='coerce')
                if pd.notnull(val) and val > 0:
                    sap_entries.append({
                        'Customer': row['SAP'], 'Document_Date': raw_date, 'Posting_Date': raw_date,
                        'Company_Code': 'MKPL', 'Currency_Rate': 'INR', 'Reference': period_val,
                        'Doc_Header_Text': period_val, 'Assignment': 'MP payout', 'GL': gl_code,
                        'GST': '', 'Amount_GL': -abs(round(float(val), 2)), 'Amount_GST': '',
                        'Amount_Customer': abs(round(float(val), 2)), 'Profit_Center': '',
                        'Cost_Center': 'FINANCE', 'Text': f"MP payout {txt_prefix} {period_val}"
                    })

        wb = load_workbook(SAP_TEMPLATE_PATH)
        ws = wb.active
        for i, entry in enumerate(sap_entries):
            for c, val in enumerate(list(entry.values()), 1):
                ws.cell(row=2 + i, column=c).value = val
        wb.save(SAP_OUTPUT_FILE)
        
        google_sap_list = ["8302253", "8302697"]
        df_google = pay_df[pay_df["SAP"].astype(str).isin(google_sap_list)].copy()
        wb_g = load_workbook(SAP_TEMPLATE_PATH)
        ws_g = wb_g.active
        for i, (_, row) in enumerate(df_google.iterrows()):
            amt = round(float(pd.to_numeric(row["GMV"], errors="coerce") or 0), 2)
            data = [row["SAP"], raw_date, raw_date, "MKPL", "INR", period_val, period_val, "MP payout", "91000467", "", -amt, "", amt, "", "FINANCE", f"MP payout Escrow {period_val}"]
            for c, v in enumerate(data, 1): ws_g.cell(row=2 + i, column=c).value = v
        wb_g.save(GOOGLE_UPLOADER_FILE)
        print("✅ SAP and Google uploaders generated.")
    except Exception as e:
        print(f"❌ SAP/Google Error: {e}")

# ---------------- 3. JV UPLOADER GENERATION ---------------- #

def generate_jv_uploaders(period_val, raw_date):
    try:
        jv_working = pd.read_excel(JV_WORKING_FILE)
        configs = [
            {'output': JV_GROSS_OUTPUT, 'keep': 'sales gross', 'rem': ['Commission', 'sales return', 'cart'], 'col': 'direction_1', 'ser': 1, 'pk_m': '40', 'pk_s': '11', 'acc': '13025016', 'txt': 'MP Payt Gross', 'm': 1},
            {'output': JV_RETURN_OUTPUT, 'keep': 'sales return', 'rem': ['Commission', 'sales gross', 'cart'], 'col': 'direction_2', 'ser': 2, 'pk_m': '50', 'pk_s': '11', 'acc': '13025014', 'txt': 'MP Refunded Gross', 'm': -1}
        ]
        for cfg in configs:
            wb = load_workbook(JV_FORMAT_FILE)
            for sheet_to_del in cfg['rem']:
                if sheet_to_del in wb.sheetnames: del wb[sheet_to_del]
                
            ws = wb[cfg['keep']]
            total = round(jv_working[cfg['col']].sum() * cfg['m'], 2)
            
            ws.cell(row=2, column=1).value = cfg['ser']
            ws.cell(row=2, column=2).value = raw_date
            ws.cell(row=2, column=3).value = "SA"
            ws.cell(row=2, column=4).value = "MKPL"
            ws.cell(row=2, column=5).value = raw_date
            ws.cell(row=2, column=6).value = "INR"
            ws.cell(row=2, column=7).value = period_val
            ws.cell(row=2, column=8).value = period_val
            ws.cell(row=2, column=9).value = cfg['pk_m']
            ws.cell(row=2, column=10).value = cfg['acc']
            ws.cell(row=2, column=13).value = total
            ws.cell(row=2, column=27).value = f"{cfg['txt']} {period_val}"
            
            curr = 3
            for _, row in jv_working.iterrows():
                val = round(float(row[cfg['col']]) * cfg['m'], 2)
                if val == 0: continue
                ws.cell(row=curr, column=1).value = cfg['ser']
                ws.cell(row=curr, column=9).value = cfg['pk_s']
                ws.cell(row=curr, column=10).value = row["SAP"]
                ws.cell(row=curr, column=13).value = val
                ws.cell(row=curr, column=27).value = f"{cfg['txt']} {period_val}"
                curr += 1
            
            for r in range(2, curr):
                for col_idx, val_static in enumerate(["v0", "1000", "1000", "0001"], 14):
                    ws.cell(row=r, column=col_idx).value = val_static
                ws.cell(row=r, column=22).value, ws.cell(row=r, column=23).value = "Gross", "FINANCE"
                
            wb.save(cfg['output'])
        
        generate_commission_uploader(jv_working, period_val, raw_date)
        generate_cart_uploader(period_val, raw_date)
    except Exception as e:
        print(f"❌ JV Generation Error: {e}")

def generate_commission_uploader(jv_working, period_val, raw_date):
    try:
        wb = load_workbook(JV_FORMAT_FILE)
        for s in ['sales gross', 'sales return', 'cart']:
            if s in wb.sheetnames: del wb[s]
        ws = wb['Commission']
        total_comm_tax = round((jv_working['Commission'] + jv_working['Tax']).sum(), 2)
        h1_amt, h2_amt = round(total_comm_tax / 1.18, 2), round(total_comm_tax - (total_comm_tax / 1.18), 2)
        
        for r, amt, gl_code in [(2, h1_amt, "83001144"), (3, h2_amt, "12000417")]:
            ws.cell(row=r, column=1).value = 3
            ws.cell(row=r, column=2).value = raw_date
            ws.cell(row=r, column=3).value = "SA"
            ws.cell(row=r, column=4).value = "MKPL"
            ws.cell(row=r, column=5).value = raw_date
            ws.cell(row=r, column=6).value = "INR"
            ws.cell(row=r, column=7).value = period_val
            ws.cell(row=r, column=8).value = period_val
            ws.cell(row=r, column=9).value = "50"
            ws.cell(row=r, column=10).value = gl_code
            ws.cell(row=r, column=13).value = amt

        curr = 4
        for _, row in jv_working.iterrows():
            amt = round(row['Commission'] + row['Tax'], 2)
            if amt == 0: continue
            ws.cell(row=curr, column=1).value = 3
            ws.cell(row=curr, column=9).value = "01"
            ws.cell(row=curr, column=10).value = row["SAP"]
            ws.cell(row=curr, column=13).value = amt
            curr += 1
            
        for r in range(2, curr):
            for col_idx, val_static in enumerate(["v0", "1000", "1000", "0001"], 14):
                ws.cell(row=r, column=col_idx).value = val_static
            ws.cell(row=r, column=22).value, ws.cell(row=r, column=23).value = "Commission", "FINANCE"
            ws.cell(row=r, column=27).value = f"MP Payt Commission {period_val}"
            
        wb.save(JV_COMM_OUTPUT)
        print("✅ Commission Uploader generated.")
    except Exception as e: print(f"❌ Commission Error: {e}")

def generate_cart_uploader(period_val, raw_date):
    try:
        output_df = pd.read_excel(OUTPUT_FILE)
        sum_shipping = output_df['shipping_amount'].sum()
        sum_reversal = output_df['cust_shipping_reversal'].sum()
        sum_cart_conv = output_df['cart_conv_fee'].sum()

        amt_13025016 = round(sum_shipping + sum_reversal + (sum_cart_conv * 1.18), 2)
        amt_83001613 = round(amt_13025016 / 1.18, 2)
        amt_12000417 = round(amt_83001613 * 0.18, 2)

        wb = load_workbook(JV_FORMAT_FILE)
        for s in ['sales gross', 'sales return', 'Commission']:
            if s in wb.sheetnames: del wb[s]
        
        ws = wb['cart']
        # --- Applied corrections: SerialNo=1, Posting_Date=DocumentDate, Reference=period_val, HeaderText=Reference ---
        ws.cell(row=2, column=1).value = 1              # SerialNo
        ws.cell(row=2, column=2).value = raw_date       # DocumentDate
        ws.cell(row=2, column=5).value = raw_date       # Posting_Date (Match DocumentDate)
        ws.cell(row=2, column=7).value = period_val     # Reference
        ws.cell(row=2, column=8).value = period_val     # HeaderText (Match Reference)
        
        ws.cell(row=2, column=9).value = "40"
        ws.cell(row=2, column=10).value = "13025016"
        ws.cell(row=2, column=13).value = amt_13025016
        
        gl_configs = [(3, amt_83001613, "83001613", "FINANCE"), (4, amt_12000417, "12000417", "")]
        for r, amt, gl_code, cc in gl_configs:
            ws.cell(row=r, column=1).value = 1          # SerialNo
            ws.cell(row=r, column=9).value = "50"
            ws.cell(row=r, column=10).value = gl_code
            ws.cell(row=r, column=13).value = amt
            ws.cell(row=r, column=23).value = cc

        for r in range(2, 5):
            for col_idx, val_static in enumerate(["v0", "1000", "1000", "0001"], 14):
                if ws.cell(row=r, column=col_idx).value is None:
                    ws.cell(row=r, column=col_idx).value = val_static
            ws.cell(row=r, column=27).value = f"MP cart_conv_fee Payout {period_val}"
            
        wb.save(JV_CART_OUTPUT)
        print("✅ jv_uploader_cart.xlsx generated with HeaderText=Reference fix.")
    except Exception as e: print(f"❌ Cart Uploader Error: {e}")

# ---------------- 4. MERGE & EMAIL ---------------- #

def merge_all_jv_files():
    try:
        print("Merging all JVs...")
        dfs = []
        for path in [JV_GROSS_OUTPUT, JV_RETURN_OUTPUT, JV_COMM_OUTPUT]:
            if os.path.exists(path): dfs.append(pd.read_excel(path))
        if dfs:
            combined = pd.concat(dfs, ignore_index=True)
            combined.to_excel(FINAL_MERGED_JV, index=False, sheet_name='Uploader')
    except Exception as e: print(f"❌ Merging Error: {e}")

def send_gmail(attachments, period_val):
    try:
        msg = MIMEMultipart()
        msg['Subject'] = f"Testing |SAP + Google + JV Payment uploader || {period_val}"
        msg['From'], msg['To'] = SENDER_EMAIL, ", ".join(RECEIVER_EMAIL)
        msg.attach(MIMEText(f"Attached are the uploaders for period: {period_val}.", 'plain'))

        for path in attachments:
            if os.path.exists(path):
                with open(path, "rb") as f:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(f.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Disposition', f"attachment; filename={os.path.basename(path)}")
                    msg.attach(part)

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, RECEIVER_EMAIL, msg.as_string())
        server.quit()
        print("📧 Email sent successfully.")
    except Exception as e: print(f"❌ Email error: {e}")

# ---------------- MAIN EXECUTION ---------------- #

def main_function():
    process_revenue_report()
    
    period_df = pd.read_excel(PERIOD_FILE_PATH, header=None)
    period_val = str(period_df.iloc[0, 0]).strip()
    raw_df = pd.read_csv(DATA_FILE, usecols=['settled_at'], encoding='latin1', nrows=1)
    raw_date = pd.to_datetime(raw_df['settled_at'].iloc[0]).strftime('%d.%m.%Y')
    
    generate_sap_and_google_uploads(period_val, raw_date)
    generate_jv_uploaders(period_val, raw_date)
    merge_all_jv_files()
    
    send_gmail([SAP_OUTPUT_FILE, GOOGLE_UPLOADER_FILE, FINAL_MERGED_JV, JV_CART_OUTPUT], period_val)

if __name__ == "__main__":
    main_function()